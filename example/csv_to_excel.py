#!/usr/bin/env python3
"""将 CSV 文件转换为格式化的 Excel 文件。

用法:
    python3 csv_to_excel.py <输入CSV文件> [输出Excel文件]

示例:
    python3 csv_to_excel.py free_week_12.csv
    python3 csv_to_excel.py free_week_12.csv output.xlsx
"""

import sys
import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def csv_to_formatted_excel(csv_file: str, excel_file: str = None):
    """将 CSV 转换为格式化的 Excel 文件。

    Args:
        csv_file: 输入的 CSV 文件路径
        excel_file: 输出的 Excel 文件路径（可选，默认为同名 .xlsx）
    """
    if excel_file is None:
        base_name = os.path.splitext(csv_file)[0]
        excel_file = f"{base_name}.xlsx"

    print(f"读取 CSV 文件: {csv_file}")
    df = pd.read_csv(csv_file, encoding='utf-8-sig')

    print(f"转换为 Excel: {excel_file}")
    df.to_excel(excel_file, index=False, engine='openpyxl')

    print("格式化 Excel（自动列宽、行高、自动换行）...")
    wb = load_workbook(excel_file)
    ws = wb.active

    # 设置所有单元格默认格式
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            if row_idx == 1 and col_idx == 1:
                cell.value = ""
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical='center',
                    horizontal='center'
                )
            elif row_idx == 1:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical='center',
                    horizontal='center'
                )
            elif col_idx == 1:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical='center',
                    horizontal='center'
                )
            else:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical='top',
                    horizontal='left'
                )

    # 调整列宽：第一列固定宽度，其他列根据内容智能设置
    for col_idx, column in enumerate(ws.columns, start=1):
        column_letter = column[0].column_letter

        if col_idx == 1:
            ws.column_dimensions[column_letter].width = 11
        else:
            max_length = 0
            for cell in column:
                if cell.value:
                    cell_str = str(cell.value)
                    lines = cell_str.replace('、', ',').split(',')
                    max_line_length = max(len(line.strip()) for line in lines) if lines else len(cell_str)
                    max_length = max(max_length, max_line_length)

            adjusted_width = min(max(max_length + 2, 30), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # 调整行高：根据内容自动设置
    for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
        max_lines = 1
        for cell in row:
            if cell.value:
                cell_str = str(cell.value)
                separators_count = cell_str.count('、') + cell_str.count(',')
                estimated_lines = max(1, (separators_count + 1) // 3)
                max_lines = max(max_lines, estimated_lines)

        ws.row_dimensions[row_idx].height = max(20, min(max_lines * 15, 90))

    wb.save(excel_file)
    print(f"✓ 成功生成: {excel_file}")
    print(f"  - 已设置自动换行")
    print(f"  - 已调整列宽")
    print(f"  - 已调整行高")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    csv_file = sys.argv[1]
    excel_file = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(csv_file):
        print(f"错误: 文件不存在 - {csv_file}")
        sys.exit(1)

    try:
        csv_to_formatted_excel(csv_file, excel_file)
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
